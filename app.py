import streamlit as st
import pandas as pd
import os
import tempfile
import io
from openpyxl import load_workbook

st.title("🔥 Tool Tổng Hợp Học Phí Của Quỳnh")

uploaded_files = st.file_uploader(
    "📁 Chọn file Excel",
    type=["xls", "xlsx"],
    accept_multiple_files="directory"
)

# =========================
# HÀM MỚI: CHECK TÊN SHEET
# =========================
def check_sheet_name(excel_wb):
    """
    Kiểm tra tên sheet có khớp với tên lớp trong ô A4 hay không
    """
    results = []

    for sheet_name in excel_wb.sheetnames:
        ws = excel_wb[sheet_name]
        header_value = ws["A4"].value

        if not header_value:
            status = "ERROR"
            message = f"Sheet {sheet_name}: Ô A4 đang trống hoặc không hợp lệ"
        else:
            header_str = str(header_value).strip()
            if sheet_name in header_str:
                status = "OK"
                message = f"Sheet {sheet_name}: Đã đúng tên lớp"
            else:
                status = "WARNING"
                message = f"Sheet {sheet_name}: Cần sửa lại tên lớp"

        results.append({
            "SheetName": sheet_name,
            "Header_A4": header_value,
            "Status": status,
            "Message": message
        })

    return results


if st.button("🚀 Xử lý dữ liệu"):

    if not uploaded_files:
        st.error("❌ Bạn chưa upload file nào!")
        st.stop()

    temp_dir = tempfile.mkdtemp()
    all_data = []
    all_sheet_checks = []

    for up_file in uploaded_files:

        # Đọc binary
        content = up_file.read()
        safe_name = os.path.basename(up_file.name)

        file_path = os.path.join(temp_dir, safe_name)
        with open(file_path, "wb") as f:
            f.write(content)

        st.write(f"🔄 Đang xử lý {safe_name}")

        ext = os.path.splitext(file_path)[1].lower()
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"

        # =========================
        # PHẦN CŨ: ĐỌC DATA BẰNG PANDAS
        # =========================
        xls = pd.ExcelFile(file_path, engine=engine)

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=None,
                engine=engine
            )

            header_row = 9
            start_row = 11

            if len(df) <= start_row:
                continue

            df_data = df.iloc[start_row:].dropna(how="all")
            if df_data.empty:
                continue

            fixed = df_data.iloc[:, :8]

            header_data = df.iloc[header_row]
            keep_idx = [
                i for i, v in enumerate(header_data)
                if pd.isna(v) and i >= 10
            ]
            keep = df_data.iloc[:, keep_idx]

            merged = pd.concat([fixed, keep], axis=1)
            merged.columns = range(merged.shape[1])
            merged["SheetName"] = sheet_name
            merged["FileName"] = safe_name

            all_data.append(merged)

        # =========================
        # PHẦN MỚI: CHECK TÊN SHEET
        # =========================
        if ext == ".xlsx":
            excel_wb = load_workbook(file_path, data_only=True)
            sheet_checks = check_sheet_name(excel_wb)

            for r in sheet_checks:
                r["FileName"] = safe_name

            all_sheet_checks.extend(sheet_checks)

    if not all_data:
        st.error("❌ Không có dữ liệu hợp lệ để tổng hợp")
        st.stop()

    # =========================
    # GỘP DỮ LIỆU
    # =========================
    final_df = pd.concat(all_data, ignore_index=True)
    check_df = pd.DataFrame(all_sheet_checks)

    st.success("🎉 Hoàn tất xử lý!")

    # =========================
    # XUẤT FILE EXCEL
    # =========================
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, sheet_name="TongHopHocPhi")
        if not check_df.empty:
            check_df.to_excel(writer, index=False, sheet_name="CheckTenSheet")

    buffer.seek(0)

    st.download_button(
        "⬇️ Tải file tổng hợp",
        data=buffer,
        file_name="TongHop_HocPhi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # =========================
    # HIỂN THỊ KẾT QUẢ CHECK
    # =========================
    if not check_df.empty:
        st.subheader("📋 Kiểm tra tên sheet & tên lớp")
        for _, r in check_df.iterrows():
            if r["Status"] == "OK":
                st.success(r["Message"])
            elif r["Status"] == "WARNING":
                st.warning(r["Message"])
            else:
                st.error(r["Message"])
